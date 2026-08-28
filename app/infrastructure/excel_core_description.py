"""Read interval-based sedimentological descriptions and bind them to core JPGs."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace

import cv2
import numpy as np

from PySide6.QtCore import QPointF
from PySide6.QtGui import QPixmap

from app.domain.facies_catalog import facies_metadata
from app.domain.lithology_attributes import LITHOLOGY_ATTRIBUTE_OPTIONS
from app.domain.models import FaciesDetection
from app.infrastructure.ml.rule_based_facies import RuleBasedFaciesDetector


EXCEL_FACIES_ATTRIBUTE_FIELDS = (
    "Энергия среды",
    "Гидродинамический режим",
    "Примечание",
    *LITHOLOGY_ATTRIBUTE_OPTIONS.keys(),
)

# These aliases occur often in field descriptions.  The import deliberately
# accepts both a formal 16-column template and a short table such as
# ``Интервал | Порода | Цвет | Описание``.
ATTRIBUTE_HEADER_ALIASES = {
    "Название породы": ("название породы", "порода", "литология", "литотип"),
    "Цвет": ("цвет", "окраска"),
    "Зернистость": ("зернистость", "размер зерна"),
    "Примесь другой фракции": ("примесь", "другая фракция"),
    "Флюидонасыщение": ("флюид", "насыщение"),
    "Цемент": ("цемент", "цементация"),
    "Слоистость": ("слоистость",),
    "Текстура": ("текстура",),
    "Биотурбация": ("биотурбация",),
    "Включения": ("включения",),
    "Органические остатки": ("органические остатки", "остатки"),
    "Пустотное пространство. Трещиноватость": ("трещин", "пустот"),
    "Степень цементации": ("степень цементации",),
    "Контакт": ("контакт",),
    "Ориентация контакта": ("ориентация контакта",),
    "Целостность": ("целостность",),
}


@dataclass(frozen=True)
class DescriptionLayer:
    well: str
    top: float
    base: float
    facies_name: str
    facies_code: str
    facies_index: str
    description: str
    attributes: dict[str, str]
    sheet: str
    row: int
    core_top: float | None = None
    core_base: float | None = None

    @property
    def label(self) -> str:
        return self.facies_code or self.facies_index or self.facies_name


@dataclass(frozen=True)
class CoreInterval:
    well: str
    top: float
    base: float

    @property
    def title(self) -> str:
        return f"Скв. {self.well} · {self.top:g}–{self.base:g} м"


def photo_interval_from_filename(path: Path) -> CoreInterval:
    """Read well name and measured depth range from a core-photo filename.

    Expected filename form: ``Р-31_3002,00-3004,96.jpg``.  The well prefix
    is deliberately permissive, while both depths must be explicit numbers.
    """
    # Accept real-world names too: ``Р-31 3002,00–3004,96 (1).jpg`` and
    # ``Р-31_3002,00-3004,96.jpg``.  The old full-match rejected both even
    # though their interval was unambiguous.
    match = re.search(
        r"(?P<well>.+?)[_\s]+(?P<top>\d+(?:[.,]\d+)?)\s*[-–—]\s*(?P<base>\d+(?:[.,]\d+)?)",
        path.stem,
    )
    if not match:
        raise ValueError("В имени JPG не найден интервал. Пример: «Скв_Кровля-Подошва.jpg».")
    well = _display_text(match.group("well")).rstrip(" _-")
    top = _as_float(match.group("top"))
    base = _as_float(match.group("base"))
    if base is not None and top is not None and base <= top:
        # A single source filename contains 3829,08-2829,33.  Recover only
        # this narrow, obvious first-digit typo; all other reverse intervals
        # remain errors instead of being guessed.
        top_integer = re.split(r"[.,]", match.group("top"), maxsplit=1)[0]
        base_integer, base_fraction = re.split(r"[.,]", match.group("base"), maxsplit=1)
        if len(top_integer) == len(base_integer) and top_integer[1:] == base_integer[1:]:
            base = _as_float(f"{top_integer[0]}{base_integer[1:]}.{base_fraction}")
    if not well or top is None or base is None or base <= top:
        raise ValueError("В имени JPG указан некорректный интервал глубин.")
    return CoreInterval(well=well, top=top, base=base)


@dataclass(frozen=True)
class ImportIssue:
    source: str
    message: str


def read_description_workbook(path: Path) -> tuple[list[DescriptionLayer], list[ImportIssue]]:
    """Read the wide, merged-header Excel form used for core descriptions."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Не установлен пакет openpyxl. Выполните обновление зависимостей Kern Analyzer.") from exc

    workbook = load_workbook(path, read_only=False, data_only=True)
    layers: list[DescriptionLayer] = []
    issues: list[ImportIssue] = []
    for sheet in workbook.worksheets:
        columns = _find_columns(sheet)
        has_pair = {"facies_top", "facies_base"}.issubset(columns)
        has_range = "facies_interval" in columns
        sheet_well = _well_from_sheet_title(sheet.title) if "well" not in columns else ""
        if not (has_pair or has_range) or ("well" not in columns and not sheet_well):
            continue
        last_well = ""
        for row in range(columns.get("data_start", 1), sheet.max_row + 1):
            top, base = _row_interval(sheet, row, columns)
            if top is None or base is None or base <= top:
                continue
            well = _display_text(_cell_value(sheet, row, columns.get("well"))) or last_well or sheet_well
            if not well:
                issues.append(ImportIssue(f"{sheet.title}!{row}", "Не указан номер скважины."))
                continue
            last_well = well
            name = _display_text(_cell_value(sheet, row, columns.get("facies_name")))
            code = _display_text(_cell_value(sheet, row, columns.get("facies_code")))
            index = _display_text(_cell_value(sheet, row, columns.get("facies_index")))
            description = _display_text(_cell_value(sheet, row, columns.get("description")))
            attributes = {
                field: value
                for field in EXCEL_FACIES_ATTRIBUTE_FIELDS
                if (value := _display_text(_cell_value(sheet, row, columns.get(f"attribute:{field}"))))
            }
            if _looks_like_column_number_row(name, code, index, description):
                continue
            # A field notebook frequently contains only an interval and free
            # text.  Keep it as a reviewable facies instead of dropping the
            # row: colour/lithology are still extracted into its card.
            if not (name or code or index):
                name = _infer_facies_name(description, attributes) or "Не задано"
                issues.append(ImportIssue(
                    f"{sheet.title}!{row}",
                    "Нет кода/индекса фации: создана метка «Не задано», проверьте её перед обучением.",
                ))
            layers.append(
                DescriptionLayer(
                    well=well,
                    top=top,
                    base=base,
                    facies_name=name,
                    facies_code=code,
                    facies_index=index,
                    description=description,
                    attributes=attributes,
                    sheet=sheet.title,
                    row=row,
                    core_top=_as_float(_cell_value(sheet, row, columns.get("core_top"))),
                    core_base=_as_float(_cell_value(sheet, row, columns.get("core_base"))),
                )
            )
    if not layers:
        raise ValueError(
            "Не найдены строки с интервалом фации. Укажите скважину, верх/низ "
            "(или один столбец с интервалом). Название, код и параметры можно оставить в свободном описании."
        )
    return sorted(layers, key=lambda item: (_well_key(item.well), item.top, item.base)), issues


def layers_for_photo(layers: list[DescriptionLayer], well: str, top: float, base: float) -> list[DescriptionLayer]:
    """Return description rows overlapping a photographed core interval."""
    requested_well = _well_key(well)
    return [
        layer
        for layer in layers
        if _well_key(layer.well) == requested_well and min(layer.base, base) - max(layer.top, top) > 0.00001
    ]


def core_intervals(layers: list[DescriptionLayer]) -> list[CoreInterval]:
    """Return unique core-sampling intervals used to bind a batch of JPGs."""
    grouped: dict[tuple[str, float, float], CoreInterval] = {}
    for layer in layers:
        top = layer.core_top if layer.core_top is not None else layer.top
        base = layer.core_base if layer.core_base is not None else layer.base
        if base <= top:
            continue
        item = CoreInterval(layer.well, top, base)
        grouped[(_well_key(item.well), round(item.top, 4), round(item.base, 4))] = item
    return sorted(grouped.values(), key=lambda item: (_well_key(item.well), item.top, item.base))


def create_depth_bound_detections(
    image_path: Path,
    pixmap: QPixmap,
    photo_top: float,
    photo_base: float,
    layers: list[DescriptionLayer],
) -> tuple[list[FaciesDetection], list[ImportIssue], list[dict[str, float]]]:
    """Turn matching depth intervals into per-column rectangular training labels."""
    if photo_base <= photo_top:
        raise ValueError("Низ интервала фотографии должен быть больше верха.")
    # ``cv2.imread`` fails on Windows paths containing Cyrillic characters
    # (for example ``C:\\Users\\Я\\...``).  Reading bytes through Python first
    # keeps the Excel + JPG import independent of the current user profile.
    try:
        image_bytes = image_path.read_bytes()
    except OSError as exc:
        raise ValueError(f"Не удалось прочитать изображение: {image_path.name}") from exc
    image = cv2.imdecode(np.frombuffer(image_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"Не удалось открыть изображение: {image_path.name}")
    columns = RuleBasedFaciesDetector._find_core_columns(image)
    if not columns:
        height, width = image.shape[:2]
        # A fallback remains usable for a single long core photograph. It is
        # intentionally restricted away from the usual caption band.
        columns = [(0, 0, width, max(1, int(height * 0.87)))]
    columns = sorted(columns, key=lambda box: (box[0], box[1]))
    source_height, source_width = image.shape[:2]
    x_scale = pixmap.width() / max(1, source_width)
    y_scale = pixmap.height() / max(1, source_height)
    total_visual_length = sum(max(1, bottom - top) for _, top, _, bottom in columns)
    cursor = photo_top
    detections: list[FaciesDetection] = []
    issues: list[ImportIssue] = []
    depth_segments: list[dict[str, float]] = []
    for left, pixel_top, right, pixel_base in columns:
        span = (photo_base - photo_top) * (pixel_base - pixel_top) / total_visual_length
        column_top, column_base = cursor, cursor + span
        cursor = column_base
        depth_segments.append({
            "left": float(left * x_scale), "top": float(pixel_top * y_scale), "right": float(right * x_scale), "bottom": float(pixel_base * y_scale),
            "depth_from": float(column_top), "depth_to": float(column_base),
        })
        for layer in layers:
            overlap_top, overlap_base = max(layer.top, column_top), min(layer.base, column_base)
            if overlap_base - overlap_top <= 0.00001:
                continue
            y0 = pixel_top + (overlap_top - column_top) / (column_base - column_top) * (pixel_base - pixel_top)
            y1 = pixel_top + (overlap_base - column_top) / (column_base - column_top) * (pixel_base - pixel_top)
            polygon = [
                QPointF(left * x_scale, y0 * y_scale),
                QPointF((right - 1) * x_scale, y0 * y_scale),
                QPointF((right - 1) * x_scale, y1 * y_scale),
                QPointF(left * x_scale, y1 * y_scale),
            ]
            attributes = _attributes_from_description(layer)
            detections.append(
                FaciesDetection(
                    label=layer.label,
                    confidence=1.0,
                    polygon=polygon,
                    attributes=attributes,
                    depth_from=overlap_top,
                    depth_to=overlap_base,
                    # The source is a human-written depth description. The
                    # geometry is automatic and remains visibly editable.
                    training_ready=True,
                )
            )
    if not detections:
        issues.append(ImportIssue(image_path.name, "Для фотографии не найдено пересекающихся интервалов фаций."))
    return detections, issues, depth_segments


def create_automatic_interval_detections(
    image_path: Path,
    pixmap: QPixmap,
    photo_top: float | None = None,
    photo_base: float | None = None,
) -> tuple[list[FaciesDetection], list[dict[str, float]]]:
    """Find persistent visual packages and create editable interval masks.

    This is intentionally classless: visual texture alone cannot truthfully
    name a geological facies.  It supplies the interval rows and masks, while
    the interpreter selects a facies in the card before using it for training.
    """
    try:
        image = cv2.imdecode(np.frombuffer(image_path.read_bytes(), dtype=np.uint8), cv2.IMREAD_COLOR)
    except OSError as exc:
        raise ValueError(f"Не удалось прочитать изображение: {image_path.name}") from exc
    if image is None:
        raise ValueError(f"Не удалось открыть изображение: {image_path.name}")

    detector = RuleBasedFaciesDetector()
    _, intervals = detector.analyse(image)
    columns = detector._find_core_columns(image)
    height, width = image.shape[:2]
    if not columns:
        columns = [(0, 0, width, max(1, int(height * 0.87)))]
    columns = sorted(columns, key=lambda item: (item[0], item[1]))
    if not intervals:
        intervals = [
            SimpleNamespace(
                polygon=((left, top), (right - 1, top), (right - 1, bottom - 1), (left, bottom - 1)),
                evidence="fallback core column",
            )
            for left, top, right, bottom in columns
        ]

    x_scale = pixmap.width() / max(1, width)
    y_scale = pixmap.height() / max(1, height)
    total_length = sum(max(1, bottom - top) for _, top, _, bottom in columns)
    cursor = photo_top if photo_top is not None and photo_base is not None and photo_base > photo_top else None
    segments: list[dict[str, float]] = []
    for left, top, right, bottom in columns:
        segment = {"left": float(left * x_scale), "top": float(top * y_scale), "right": float(right * x_scale), "bottom": float(bottom * y_scale)}
        if cursor is not None:
            span = (photo_base - photo_top) * (bottom - top) / total_length
            segment.update({"depth_from": float(cursor), "depth_to": float(cursor + span)})
            cursor += span
        segments.append(segment)

    detections: list[FaciesDetection] = []
    for interval in intervals:
        polygon = [QPointF(x * x_scale, y * y_scale) for x, y in interval.polygon]
        depth_from = depth_to = None
        for segment in segments:
            if segment["left"] <= polygon[0].x() <= segment["right"] and "depth_from" in segment:
                pixel_span = segment["bottom"] - segment["top"]
                depth_span = segment["depth_to"] - segment["depth_from"]
                depth_from = segment["depth_from"] + (min(point.y() for point in polygon) - segment["top"]) / pixel_span * depth_span
                depth_to = segment["depth_from"] + (max(point.y() for point in polygon) - segment["top"]) / pixel_span * depth_span
                break
        detections.append(FaciesDetection(
            label="Новый контур",
            confidence=0.65,
            polygon=polygon,
            attributes={"Источник интервала": "Автоматически по текстуре фото", "Основание": str(interval.evidence)},
            depth_from=round(depth_from, 3) if depth_from is not None else None,
            depth_to=round(depth_to, 3) if depth_to is not None else None,
        ))
    return detections, segments


def _find_columns(sheet) -> dict[str, int]:
    """Find description fields by meaning instead of a fixed Excel template.

    Excel descriptions in the field commonly have two or three header rows,
    merged group labels and slightly different wording.  We combine only the
    actual header rows, then score each column for every required role.  This
    keeps "Кровля" of the *facies* interval separate from core/GIS intervals.
    """
    header_end = _header_end_row(sheet)
    descriptions = [
        (
            column,
            " ".join(
                _normalized(_merged_value(sheet, row, column))
                for row in range(1, header_end + 1)
                if _merged_value(sheet, row, column) is not None
            ),
        )
        for column in range(1, sheet.max_column + 1)
    ]
    columns: dict[str, int] = {"data_start": header_end + 1}
    for role in ("well", "facies_name", "facies_code", "facies_index", "description", "facies_top", "facies_base", "core_top", "core_base"):
        if column := _best_column(descriptions, role):
            columns[role] = column
    # A number of exports put both limits in one cell: "3915,00–3915,55".
    # Use it only when a top/base pair was not identified.
    if not {"facies_top", "facies_base"}.issubset(columns):
        if column := _best_column(descriptions, "facies_interval"):
            columns["facies_interval"] = column
    for field in EXCEL_FACIES_ATTRIBUTE_FIELDS:
        needles = (_normalized(field), *(_normalized(value) for value in ATTRIBUTE_HEADER_ALIASES.get(field, ())))
        for column, text in descriptions:
            if any(needle in text for needle in needles):
                columns[f"attribute:{field}"] = column
                break
    return columns


def _header_end_row(sheet) -> int:
    """Return the final header row without treating the data area as a header."""
    limit = min(30, sheet.max_row)
    header_words = ("скваж", "скв", "интервал", "кровл", "подошв", "глубин", "описан", "depth", "well", "from", "to", "индекс", "код", "название")
    last = 1
    for row in range(1, limit + 1):
        text = " ".join(_normalized(_merged_value(sheet, row, column)) for column in range(1, sheet.max_column + 1))
        if any(word in text for word in header_words):
            last = row
    return last


def _best_column(descriptions: list[tuple[int, str]], role: str) -> int | None:
    scored = [(column, _column_score(text, role)) for column, text in descriptions]
    column, score = max(scored, key=lambda item: item[1], default=(0, 0))
    return column if score > 0 else None


def _column_score(text: str, role: str) -> int:
    def has(*needles: str) -> bool:
        return any(needle in text for needle in needles)

    facies_group = 90 if has("интервал фаци", "facies interval", "facies depth") else 0
    generic_interval = 35 if has("интервал", "глубин", "depth") else 0
    core_group = 70 if has("отбор керна", "core interval", "керн") else 0
    direction_top = 30 if has("кровл", "верх", "от", "from", "top") else 0
    direction_base = 30 if has("подошв", "низ", "до", "to", "base", "bottom") else 0
    if role == "well":
        return 100 if has("скваж", "№ скв", "no скв", "well") else 0
    if role == "facies_name":
        return 100 if has("название фаци", "наименование фаци", "facies name") else (55 if has("литофаци", "фация") and not has("индекс", "код", "ассоциац") else (35 if has("название") else 0))
    if role == "facies_code":
        return 100 if has("код фаци", "facies code") or ("фациальн" in text and "код" in text) else 0
    if role == "facies_index":
        return 100 if has("индекс фаци", "facies index") or ("фациальн" in text and "индекс" in text) else 0
    if role == "description":
        return 100 if has("краткое описание") else (70 if has("описан", "характерист", "description") else 0)
    if role == "facies_top":
        return facies_group + direction_top if facies_group and direction_top else (generic_interval + direction_top if direction_top else (18 if direction_top else 0))
    if role == "facies_base":
        return facies_group + direction_base if facies_group and direction_base else (generic_interval + direction_base if direction_base else (18 if direction_base else 0))
    if role == "core_top":
        return core_group + direction_top if core_group and direction_top else 0
    if role == "core_base":
        return core_group + direction_base if core_group and direction_base else 0
    if role == "facies_interval":
        return facies_group or generic_interval
    return 0


def _merged_value(sheet, row: int, column: int):
    value = sheet.cell(row, column).value
    if value is not None:
        return value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col).value
    return None


def _cell_value(sheet, row: int, column: int | None):
    return _merged_value(sheet, row, column) if column is not None else None


def _row_interval(sheet, row: int, columns: dict[str, int]) -> tuple[float | None, float | None]:
    top = _as_float(_cell_value(sheet, row, columns.get("facies_top")))
    base = _as_float(_cell_value(sheet, row, columns.get("facies_base")))
    if top is not None and base is not None:
        return top, base
    value = _cell_value(sheet, row, columns.get("facies_interval"))
    return _as_interval(value)


def _as_interval(value) -> tuple[float | None, float | None]:
    """Read a pair such as ``3915,00–3915,55`` from one Excel cell."""
    text = _display_text(value).replace("−", "-").replace("–", "-").replace("—", "-")
    # The hyphen between two positive depths is a separator, not a minus sign.
    # (Negative depths are still handled by separate top/base columns.)
    values = re.findall(r"\d+(?:[.,]\d+)?", text)
    if len(values) < 2:
        return None, None
    return _as_float(values[0]), _as_float(values[1])


def _looks_like_column_number_row(*values: str) -> bool:
    """Ignore form rows like 1, 2, …, 23 printed beneath a header."""
    present = [value for value in values if value]
    return bool(present) and all(re.fullmatch(r"\d{1,2}", value) for value in present)


def _well_from_sheet_title(title: str) -> str:
    """Use a sheet title only when it plausibly identifies a well."""
    text = _display_text(title)
    if not text or _normalized(text) in {"лист1", "sheet1", "sheet", "данные"}:
        return ""
    return text


def _attributes_from_description(layer: DescriptionLayer) -> dict[str, str]:
    text = _normalized(layer.description)
    attributes = facies_metadata(layer.facies_code)
    imported = {
        "Название фации": layer.facies_name,
        "Код фации": layer.facies_code,
        "Индекс фации": layer.facies_index,
        "Краткое описание": layer.description,
        "Источник описания": f"Excel: {layer.sheet}, строка {layer.row}",
    }
    attributes.update({key: value for key, value in imported.items() if value})
    # A prose cell like "Цвет: серый; порода: песчаник" is accepted in
    # addition to the controlled vocabulary below.  Explicit table columns
    # still win, so prepared datasets remain deterministic.
    attributes.update(_inline_description_attributes(layer.description))
    for field, options in LITHOLOGY_ATTRIBUTE_OPTIONS.items():
        matches = [option for option in options if _option_is_mentioned(text, option)]
        if len(matches) == 1:
            attributes[field] = matches[0]
        elif len(matches) > 1:
            attributes[f"{field} (несколько)"] = " | ".join(matches)
    # Explicit Excel fields are authoritative over text extraction and catalog
    # defaults. This makes the import deterministic for prepared datasets.
    attributes.update(layer.attributes)
    return {key: value for key, value in attributes.items() if value}


def _inline_description_attributes(description: str) -> dict[str, str]:
    """Read short ``field: value`` fragments from a free-text description."""
    result: dict[str, str] = {}
    for field, aliases in ATTRIBUTE_HEADER_ALIASES.items():
        for alias in aliases:
            match = re.search(
                rf"(?:^|[;,.\n])\s*{re.escape(alias)}\s*[:=\-]\s*([^;,.\n]+)",
                description,
                flags=re.IGNORECASE,
            )
            if match:
                result[field] = _display_text(match.group(1))
                break
    return result


def _infer_facies_name(description: str, attributes: dict[str, str]) -> str:
    """Give an otherwise unlabelled row a stable, visible review label."""
    for pattern in (r"(?:фация|литофация)\s*[:=\-]\s*([^;,.\n]+)", r"(?:название)\s*[:=\-]\s*([^;,.\n]+)"):
        match = re.search(pattern, description, flags=re.IGNORECASE)
        if match:
            return _display_text(match.group(1))
    return _display_text(attributes.get("Название породы"))


def _option_is_mentioned(text: str, option: str) -> bool:
    value = _normalized(option)
    if value in text:
        return True
    # Russian descriptions frequently use an inflected form: «алевролитов»
    # instead of «алевролит». Matching a sufficiently long root is safer than
    # an unrestricted substring and leaves ambiguous cases for review.
    for word in re.findall(r"[a-zа-яё]+", value):
        if len(word) >= 7 and word[:6] in text:
            return True
    return False


def _as_float(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    candidate = str(value).strip().replace(",", ".").replace(" ", "")
    match = re.search(r"-?\d+(?:\.\d+)?", candidate)
    return float(match.group()) if match else None


def _display_text(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def _normalized(value) -> str:
    return " ".join(
        unicodedata.normalize("NFKC", _display_text(value)).casefold().replace("ё", "е").split()
    )


def _well_key(value: str) -> str:
    text = _normalized(value)
    text = re.sub(r"\b(?:скважина|скв|well)\b", "", text)
    return re.sub(r"[^a-zа-я0-9]+", "", text)
