"""End-to-end test of Kern Analyzer's no-dialog standard pipeline."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook

from app.infrastructure.kern_analyzer_pipeline import KernAnalyzerAutomaticPipeline, _load_column_manifest


class KernAnalyzerPipelineTests(unittest.TestCase):
    def test_reuses_relative_manifest_after_drive_letter_changes(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "photos"
            source.mkdir()
            image_path = source / "Р-31_1000,00-1001,00.jpg"
            image_path.write_bytes(b"test")
            review = source / "_core_tape_review"
            review.mkdir()
            (review / "manifest.json").write_text(json.dumps({
                "source_folder": r"D:\\Керн фото\\5\\По интервалам\\Объед",
                "columns": [{"source_relative": image_path.name, "source_rectangle_px": {"left": 30, "top": 10, "right": 110, "bottom": 190}}],
            }), encoding="utf-8")

            known = _load_column_manifest(source)

            self.assertEqual([(30, 10, 110, 190)], known[image_path.name])

    def test_creates_excel_bound_masks_without_dialogs(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "photos"
            source.mkdir()
            image_path = source / "P-31_1000,00-1002,00.jpg"
            image = np.full((300, 200, 3), 240, dtype=np.uint8)
            image[20:280, 60:140] = 80
            ok, encoded = cv2.imencode(".jpg", image)
            self.assertTrue(ok)
            encoded.tofile(str(image_path))

            review = source / "_core_tape_review"
            review.mkdir()
            (review / "manifest.json").write_text(json.dumps({
                "source_folder": str(source.resolve()),
                "columns": [{"source_relative": image_path.name, "source_rectangle_px": {"left": 60, "top": 20, "right": 140, "bottom": 280}}],
            }), encoding="utf-8")

            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "№ скв."
            sheet.merge_cells("A1:A2")
            sheet["C1"] = "Интервал фации по бурению, м"
            sheet.merge_cells("C1:D1")
            sheet["C2"] = "Кровля"
            sheet["D2"] = "Подошва"
            sheet["E1"] = "Название фации"
            sheet.merge_cells("E1:E2")
            sheet.append(["P-31", None, 1000.0, 1001.0, "Фация А"])
            sheet.append([None, None, 1001.0, 1002.0, "Фация Б"])
            excel = source / "description.xlsx"
            workbook.save(excel)

            result = KernAnalyzerAutomaticPipeline().run(source, excel, Path(temporary) / "result")

            self.assertEqual(1, result.photos_labeled)
            self.assertEqual(2, result.masks_created)
            self.assertEqual(["Фация А", "Фация Б"], result.classes)
            self.assertTrue((result.output_dir / "review.html").is_file())
            self.assertTrue((result.output_dir / "yolo_labels" / "0001_P-31_1000,00-1002,00.txt").is_file())

    def test_uses_folder_order_and_excel_range_when_filename_has_no_depth(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "photos"
            source.mkdir()
            image_path = source / "page_001.jpg"
            image = np.full((200, 140, 3), 230, dtype=np.uint8)
            ok, encoded = cv2.imencode(".jpg", image)
            self.assertTrue(ok)
            encoded.tofile(str(image_path))
            review = source / "_core_tape_review"
            review.mkdir()
            (review / "manifest.json").write_text(json.dumps({
                "source_folder": str(source.resolve()),
                "columns": [{"source_relative": image_path.name, "source_rectangle_px": {"left": 30, "top": 10, "right": 110, "bottom": 190}}],
            }), encoding="utf-8")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["№ скв.", "Интервал фации по бурению, м кровля", "Интервал фации по бурению, м подошва", "Название фации"])
            sheet.append(["P-31", 1200.0, 1201.0, "Фация А"])
            excel = source / "description.xlsx"
            workbook.save(excel)

            result = KernAnalyzerAutomaticPipeline().run(source, excel, Path(temporary) / "result")

            self.assertEqual(1, result.photos_labeled)
            self.assertEqual(1, result.masks_created)
            self.assertTrue(any("общая привязка" in item.message for item in result.issues))

    def test_ignores_generated_masks_from_a_previous_run(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "photos"
            source.mkdir()
            image_path = source / "Р-31_1000,00-1001,00.jpg"
            image = np.full((200, 140, 3), 230, dtype=np.uint8)
            image[10:190, 30:110] = 80
            ok, encoded = cv2.imencode(".jpg", image)
            self.assertTrue(ok)
            encoded.tofile(str(image_path))
            old_output = source / "_kern_analyzer_previous" / "class_masks"
            old_output.mkdir(parents=True)
            encoded.tofile(str(old_output / "mask.png"))
            review = source / "_core_tape_review"
            review.mkdir()
            (review / "manifest.json").write_text(json.dumps({
                "source_folder": str(source.resolve()),
                "columns": [{"source_relative": image_path.name, "source_rectangle_px": {"left": 30, "top": 10, "right": 110, "bottom": 190}}],
            }), encoding="utf-8")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["№ скв.", "Интервал фации по бурению, м кровля", "Интервал фации по бурению, м подошва", "Название фации"])
            sheet.append(["Р-31", 1000.0, 1001.0, "Фация А"])
            excel = source / "description.xlsx"
            workbook.save(excel)

            result = KernAnalyzerAutomaticPipeline().run(source, excel, Path(temporary) / "result")

            self.assertEqual(1, result.photos_seen)
            self.assertEqual(1, result.photos_labeled)


if __name__ == "__main__":
    unittest.main()
