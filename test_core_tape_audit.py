"""Regression test for the portable Excel + core-tape audit."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from build_core_tape import run_audit


class CoreTapeAuditTests(unittest.TestCase):
    def test_audit_sums_drilling_and_photo_covered_meters(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary) / "dataset"
            well_folder = root / "well"
            tape_folder = well_folder / "_core_tape_review"
            tape_folder.mkdir(parents=True)
            (tape_folder / "manifest.json").write_text(json.dumps({
                "photos": [
                    {"source_relative": "P-31_1000,00-1001,00.jpg"},
                    {"source_relative": "P-31_1001,00-1001,40.jpg"},
                ],
            }), encoding="utf-8")

            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "№ скв."
            sheet.merge_cells("A1:A2")
            sheet["C1"] = "Интервал фации по бурению, м"
            sheet.merge_cells("C1:D1")
            sheet["C2"] = "Кровля"
            sheet["D2"] = "Подошва"
            sheet["E1"] = "Название фации"
            sheet.merge_cells("E1:E2")
            sheet.append(["P-31", None, 1000.0, 1000.5, "Фация А"])
            sheet.append([None, None, 1000.5, 1001.4, "Фация А"])
            workbook.save(well_folder / "description.xlsx")

            result = run_audit(root, root / "_facies_audit", [])

            self.assertEqual(0, result)
            report = (root / "_facies_audit" / "facies_summary.csv").read_text(encoding="utf-8-sig")
            self.assertIn("Фация А", report)
            self.assertIn("1.4", report)
            self.assertTrue((root / "_facies_audit" / "facies_audit.html").is_file())


if __name__ == "__main__":
    unittest.main()
