"""Regression tests for flexible Excel-to-facies mapping."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.infrastructure.excel_core_description import (
    _attributes_from_description,
    photo_interval_from_filename,
    read_description_workbook,
)


class ExcelDescriptionImportTests(unittest.TestCase):
    def _read(self, workbook: Workbook):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "description.xlsx"
            workbook.save(path)
            return read_description_workbook(path)

    def test_reads_merged_multirow_facies_interval_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Описание"
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:D1")
        sheet.merge_cells("E1:E2")
        sheet.merge_cells("F1:F2")
        sheet.merge_cells("G1:G2")
        sheet["A1"] = "№ скв."
        sheet["B1"] = "Интервал фации по бурению, м"
        sheet["C1"] = "Интервал фации по бурению, м"
        sheet["C2"] = "Кровля"
        sheet["D2"] = "Подошва"
        sheet["E1"] = "Индекс фации"
        sheet["F1"] = "Название фации"
        sheet["G1"] = "Краткое описание"
        sheet.append(["Р-31", None, 3915.00, 3915.55, "Shelf", "Отложения шельфа", "Алевролит"])
        sheet.append([None, None, 3915.55, 3915.77, "TL", "Трансгрессивный слой", "Песчаник"])

        layers, issues = self._read(workbook)

        self.assertEqual([], issues)
        self.assertEqual(2, len(layers))
        self.assertEqual("Р-31", layers[1].well)
        self.assertEqual("TL", layers[1].facies_index)
        self.assertAlmostEqual(3915.55, layers[1].top)

    def test_reads_generic_from_to_columns_and_sheet_well(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Скв. Р-31"
        sheet.append(["Глубина от, м", "Глубина до, м", "Фациальный индекс", "Литофация", "Характеристика"])
        sheet.append([3002.0, 3002.45, "L", "Лагуна", "Тонкозернистый песчаник"])

        layers, issues = self._read(workbook)

        self.assertEqual([], issues)
        self.assertEqual(1, len(layers))
        self.assertEqual("Скв. Р-31", layers[0].well)
        self.assertEqual("L", layers[0].facies_index)
        self.assertEqual("Лагуна", layers[0].facies_name)

    def test_reads_single_range_column_and_does_not_import_header_numbers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Скважина", "Интервал фации, м", "Код фации", "Описание"])
        sheet.append([1, 2, 3, 4])
        sheet.append(["12", "1800,0–1800,5", "A", "Глинистый алевролит"])

        layers, issues = self._read(workbook)

        self.assertEqual([], issues)
        self.assertEqual(1, len(layers))
        self.assertEqual("A", layers[0].facies_code)
        self.assertAlmostEqual(1800.5, layers[0].base)

    def test_reads_short_free_form_description_and_common_attribute_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Р-41"
        sheet.append(["Верх", "Низ", "Порода", "Окраска", "Описание"])
        sheet.append([2500.0, 2500.4, "Песчаник", "Серый", "Цемент: карбонатный; слоистость: волнистая"])

        layers, issues = self._read(workbook)

        self.assertEqual(1, len(layers))
        self.assertEqual("Песчаник", layers[0].facies_name)
        self.assertTrue(issues)
        attributes = _attributes_from_description(layers[0])
        self.assertEqual("Песчаник", attributes["Название породы"])
        self.assertEqual("Серый", attributes["Цвет"])
        self.assertEqual("карбонатный", attributes["Цемент"])

    def test_reads_interval_from_flexible_photo_filename(self):
        interval = photo_interval_from_filename(Path("Р-31 3002,00–3004,96 (1).jpg"))

        self.assertEqual("Р-31", interval.well)
        self.assertAlmostEqual(3002.0, interval.top)
        self.assertAlmostEqual(3004.96, interval.base)


if __name__ == "__main__":
    unittest.main()
