"""Regression tests for the compact editable Excel layer-description form."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.infrastructure.core_report_export import OFFICIAL_HEADERS, export_core_description_report


class CoreReportExportTests(unittest.TestCase):
    def test_lithology_guide_is_written_into_one_final_description_cell(self):
        wells = [{
            "name": "Р-31",
            "core_interval": (3002.0, 3004.0),
            "layers": [{
                "label": "L",
                "depth_from": 3002.0,
                "depth_to": 3002.5,
                "attributes": {
                    "Название фации": "Лагуна",
                    "Название породы": "Песчаник",
                    "Цвет": "Серый",
                    "Цемент": "Карбонатный",
                },
            }],
        }]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            export_core_description_report(path, "Тест", wells, "xlsx")
            sheet = load_workbook(path, data_only=True).active

            self.assertEqual(18, len(OFFICIAL_HEADERS))
            self.assertEqual("Литологическое описание (16 параметров)", sheet.cell(3, 18).value)
            self.assertEqual("Название породы: Песчаник\nЦвет: Серый\nЦемент: Карбонатный", sheet.cell(5, 18).value)
            self.assertIsNone(sheet.cell(5, 19).value)


if __name__ == "__main__":
    unittest.main()
